#!/usr/bin/env python3
"""
DSPG Curriculum Book Generator
================================
Usage:
    python generate_curriculum.py                          # uses defaults
    python generate_curriculum.py my_curriculum.xlsx       # custom Excel path
    python generate_curriculum.py input.xlsx output.html   # custom both

The script reads the Excel workbook (Master sheet + Week 1–10 sheets) and
produces a self-contained HTML curriculum book.

Excel format expected
---------------------
Master sheet  — columns: [blank], Topics, Important Milestones, Deliverables
                rows:     Week 1 … Week 10

Week N sheets — columns: Date, Event/Activities, Time, Location, Who/Responsible Parties, Deliverables
  • Rows where col A is non-empty are day headers (e.g. "Mon, Jun 1")
  • Rows where col A is None are sub-events under that day
  • Rows where col A starts with "Deliverable" are week-level deliverables
  • A day with text like "holiday", "no activities", "full day" is flagged as holiday
  • Sheets with only topic-level rows (no Date column) are rendered as topic lists

Weeks 4–10 (sparse sheets) — if the sheet has no Date column rows, the script
renders whatever rows exist as a simple topic/focus list.
"""

import sys
import re
import datetime
from pathlib import Path
from html import escape

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
DEFAULT_XLSX = "curriculum-2026.xlsx"
DEFAULT_HTML = "dspg_curriculum_2026.html"
TITLE        = "DSPG Curriculum"
SUBTITLE     = "Data Science for the Public Good · Virginia Tech"
DATE_RANGE   = "May – August 2026"

# Keywords that mark a day as a holiday / no-work day
HOLIDAY_KEYWORDS = ["holiday", "no activities",
                    "memorial day", "independence day", "labor day", "thanksgiving"]

# Keywords that hint a row is a Workshop (→ purple card)
WORKSHOP_KEYWORDS = ["workshop", "lecture", "seminar", "training"]

# Keywords that hint a row is a Deliverable (→ green card)
DELIVERABLE_KEYWORDS = ["deliverable", "deadline", "submit", "abstract", "deploy"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def h(text):
    """HTML-escape a string, preserve None as empty."""
    if text is None:
        return ""
    return escape(str(text).strip())

def is_holiday(text):
    if not text:
        return False
    t = str(text).lower()
    return any(k in t for k in HOLIDAY_KEYWORDS)

def card_color(activity, who):
    """Pick ev-blue / ev-purple / ev-green based on row content."""
    combined = ((activity or "") + " " + (who or "")).lower()
    if any(k in combined for k in DELIVERABLE_KEYWORDS):
        return "ev-green"
    if any(k in combined for k in WORKSHOP_KEYWORDS):
        return "ev-purple"
    return "ev-blue"

def split_label(label):
    """
    Split a label like "Workshop · 9:00-10:00 AM · Dr. Le Wang"
    into (time_str, who_str).
    """
    if not label:
        return "", ""
    parts = [p.strip() for p in re.split(r"·", label)]
    WHO_HINTS = ["dr.", "drs.", "fellow", "faculty", "team", "everyone",
                 "zhao", "jesse", "nathaniel", "gao", "wang", "cary",
                 "kaufman", "dahal", "cheng", "yujuan", "clancy", "serrano",
                 "junghwan", "shao", "yuanyuan", "huaiyang", "stakeholder",
                 "all teams", "pragati"]
    TIME_HINTS = ["am", "pm", "hr", "hour", "scheduled", "afternoon",
                  "morning", "continuation", "optional"]
    time_parts, who_parts = [], []
    for i, p in enumerate(parts):
        pl = p.lower()
        is_who  = any(k in pl for k in WHO_HINTS)
        is_time = any(k in pl for k in TIME_HINTS)
        if is_who:
            who_parts.append(p)
        elif i == 0 and not is_who:
            time_parts.append(p)
        elif is_time:
            time_parts.append(p)
        else:
            time_parts.append(p)
    return " · ".join(time_parts), " · ".join(who_parts)

def fmt_time(val):
    """Format a datetime.time or string time value."""
    if val is None:
        return ""
    if isinstance(val, datetime.time):
        hour = val.hour
        minute = val.minute
        ampm = "AM" if hour < 12 else "PM"
        h12 = hour % 12 or 12
        return f"{h12}:{minute:02d} {ampm}"
    return str(val).strip()

def ev_card_html(activity, time_val, who, color=None, delay=0):
    """Render a single event card."""
    act_text = h(activity)
    time_text = h(fmt_time(time_val)) if time_val else ""
    who_text  = h(who) if who else ""
    clr = color or card_color(activity, who)

    meta_html = ""
    if time_text or who_text:
        meta_html = '<div class="ev-meta">'
        if time_text:
            meta_html += f'<span class="ev-time">{time_text}</span>'
        if who_text:
            meta_html += f'<span class="ev-who">{who_text}</span>'
        meta_html += "</div>"

    style = f' style="animation-delay:{delay:.2f}s"' if delay else ""
    return (f'<div class="ev-card {clr}"{style}>'
            f'<div class="ev-bar"></div>'
            f'<div class="ev-content">{meta_html}'
            f'<div class="ev-value">{act_text}</div>'
            f'</div></div>\n')

# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_master(ws):
    """Return list of dicts: week_num, topic, milestone, deliverable."""
    weeks = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        label = str(row[0]).strip()
        if not label.lower().startswith("week"):
            continue
        num_match = re.search(r"\d+", label)
        num = int(num_match.group()) if num_match else len(weeks) + 1
        weeks.append({
            "num":        num,
            "topic":      str(row[1]).strip() if row[1] else "",
            "milestone":  str(row[2]).strip() if row[2] else "",
            "deliverable": str(row[3]).strip() if row[3] else "",
        })
    return weeks


def parse_week_sheet(ws):
    """
    Parse a week sheet into a structured dict.
    Returns:
      {
        "title": str,
        "has_days": bool,
        "days": [ { "date": str, "title": str, "holiday": bool, "events": [...], "deliverables": [...] } ],
        "topics": [ { "title": str, "instructor": str, "time": str, "note": str } ],
        "week_deliverables": [str],
      }
    """
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            rows.append(row)

    if not rows:
        return {"title": "", "has_days": False, "days": [], "topics": [], "week_deliverables": []}

    # First row is usually the sheet title
    sheet_title = str(rows[0][0]).strip() if rows[0][0] else ""

    # Detect if sheet has a proper day-column structure
    # A "Date" header row signals structured layout
    has_days = False
    for row in rows[:5]:
        if row[0] and str(row[0]).strip().lower() == "date":
            has_days = True
            break
    # Also detect if first-column values look like dates (Mon/Tue/Wed etc.)
    if not has_days:
        for row in rows[1:6]:
            if row[0] and re.match(r"(Mon|Tue|Wed|Thu|Fri|Sat|Sun)", str(row[0]).strip()):
                has_days = True
                break

    days = []
    topics = []
    week_deliverables = []

    if has_days:
        current_day = None
        in_deliverables = False
        for row in rows[1:]:  # skip title row
            col0 = str(row[0]).strip() if row[0] is not None else ""
            col1 = str(row[1]).strip() if row[1] is not None else ""
            col2 = fmt_time(row[2]) if row[2] is not None else ""
            col4 = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

            # Skip header row
            if col0.lower() == "date":
                continue

            # Week-level deliverable rows
            if "deliverable" in col0.lower():
                in_deliverables = True
                if col1:
                    week_deliverables.append(col1)
                continue

            # Continuation of deliverables (col0 empty, in_deliverables mode)
            if in_deliverables and col0 == "" and col1:
                week_deliverables.append(col1)
                continue

            # Day header row (col0 is non-empty and looks like a date)
            if col0 and re.match(r"(Mon|Tue|Wed|Thu|Fri|Sat|Sun|Sun)", col0):
                current_day = {
                    "date":    col0,
                    "title":   col1 if col1 else col0,
                    "holiday": is_holiday(col1) or is_holiday(col0),
                    "events":  [],
                    "deliverables": [],
                }
                days.append(current_day)
                continue

            # Sub-event row (col0 is empty)
            if col0 == "" and current_day is not None:
                if col1 in ("Morning:", "Afternoon:", "Evening:"):
                    continue  # section dividers — skip, time already in events
                if col1:
                    clr = card_color(col1, col4)
                    current_day["events"].append({
                        "activity": col1,
                        "time":     col2,
                        "who":      col4,
                        "color":    clr,
                    })
    else:
        # Sparse sheet — treat rows as topic list
        for row in rows[1:]:
            title = str(row[0]).strip() if row[0] is not None else ""
            inst  = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            note  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            timing= str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            if not title and not inst:
                continue
            if title.lower().startswith("week"):
                continue
            # Sometimes instructor is in col1 with timing in col3
            topics.append({
                "title":      title,
                "instructor": inst,
                "time":       timing,
                "note":       note,
            })

    return {
        "title":             sheet_title,
        "has_days":          has_days,
        "days":              days,
        "topics":            topics,
        "week_deliverables": week_deliverables,
    }

# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

CSS = """
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#fff;--page:#f8f9fa;--page-edge:#edf0f3;--border:#e2e5e9;--border-light:#ced4da;
  --text:#1a1a2e;--text-dim:#495057;--text-muted:#868e96;
  --accent:#2b5797;--accent-glow:rgba(43,87,151,0.06);
  --blue:#2563eb;--blue-bg:rgba(37,99,235,0.05);--blue-border:rgba(37,99,235,0.15);
  --green:#16a34a;--green-bg:rgba(22,163,74,0.05);--green-border:rgba(22,163,74,0.15);
  --purple:#7c3aed;--purple-bg:rgba(124,58,237,0.05);--purple-border:rgba(124,58,237,0.15);
  --deadline:#dc2626;--holiday:#d97706;--holiday-bg:rgba(217,119,6,0.06)
}
body{font-family:'JetBrains Mono',monospace;background:var(--bg);color:var(--text);min-height:100vh;overflow-x:hidden}
.book{max-width:920px;margin:0 auto;padding:3rem 2rem 4rem}
.crumbs{display:flex;align-items:center;gap:0.4rem;margin-bottom:1.8rem;font-size:0.62rem;flex-wrap:wrap}
.crumb{color:var(--accent);cursor:pointer;transition:color 0.2s}.crumb:hover{color:var(--text)}
.crumb-sep{color:var(--text-muted)}.crumb-current{color:var(--text-dim)}
.page{display:none;animation:pageIn 0.4s ease both}.page.active{display:block}
@keyframes pageIn{from{opacity:0;transform:translateX(30px)}to{opacity:1;transform:translateX(0)}}
.cover{text-align:center;padding:2rem 0 2.5rem;border-bottom:1px solid var(--border);margin-bottom:2rem}
.cover-orn{font-size:1.4rem;color:var(--accent);letter-spacing:0.5em;margin-bottom:1rem}
.cover h1{font-family:'Cormorant Garamond',serif;font-weight:700;font-size:2.4rem;letter-spacing:-0.02em;color:var(--text);line-height:1.15;margin-bottom:0.5rem}
.cover .sub{font-family:'Cormorant Garamond',serif;font-style:italic;font-size:1rem;color:var(--text-dim);margin-bottom:1rem}
.cover .dates{font-size:0.65rem;color:var(--text-muted);letter-spacing:0.15em;text-transform:uppercase}
.toc-stats{display:flex;justify-content:center;gap:2.5rem;margin-bottom:2rem;flex-wrap:wrap}
.ts{text-align:center}.ts-val{font-family:'Cormorant Garamond',serif;font-weight:700;font-size:1.8rem;color:var(--accent)}
.ts-lbl{font-size:0.58rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.12em}
.toc-heading{font-family:'Cormorant Garamond',serif;font-weight:600;font-size:0.85rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.2em;margin-bottom:1.2rem;text-align:center}
.chapter{display:flex;justify-content:space-between;align-items:center;padding:1.1rem 1.4rem;margin-bottom:0.5rem;background:var(--page);border:1px solid var(--border);border-radius:8px;cursor:pointer;transition:all 0.25s ease;animation:fadeIn 0.5s ease both;position:relative;overflow:hidden}
.chapter::before{content:'';position:absolute;left:0;top:0;bottom:0;width:0;background:var(--accent);transition:width 0.25s}
.chapter:hover{border-color:var(--accent);background:var(--page-edge);transform:translateX(4px);box-shadow:0 2px 12px rgba(0,0,0,0.06)}.chapter:hover::before{width:3px}.chapter:hover .ch-arrow{opacity:1;transform:translateX(0)}
@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
.ch-left{flex:1}.ch-num{font-size:0.58rem;display:inline-block;background:var(--accent-glow);border:1px solid var(--accent-border);color:var(--accent);text-transform:uppercase;letter-spacing:0.15em;margin-bottom:0.35rem;font-weight:600;padding:0.1rem 0.5rem;border-radius:6px}
.ch-title{font-family:'Cormorant Garamond',serif;font-weight:700;font-size:1.15rem;color:var(--text);margin-bottom:0.3rem;line-height:1.3}
.ch-meta{font-size:0.58rem;color:var(--text-muted);display:flex;gap:0.5rem;align-items:center;flex-wrap:wrap}
.ch-deliv{font-size:0.52rem;padding:0.1rem 0.45rem;border-radius:8px;background:var(--green-bg);border:1px solid var(--green-border);color:var(--green)}
.ch-milestone{font-size:0.52rem;padding:0.1rem 0.45rem;border-radius:8px;background:var(--blue-bg);border:1px solid var(--blue-border);color:var(--blue)}
.ch-right{display:flex;align-items:center;margin-left:1rem}
.ch-arrow{font-size:1.1rem;color:var(--accent);opacity:0;transform:translateX(-6px);transition:all 0.25s}
.wp-header{margin-bottom:1.5rem}.wp-label{font-size:0.65rem;color:var(--accent);text-transform:uppercase;letter-spacing:0.2em;margin-bottom:0.4rem;font-weight:500}
.wp-title{font-family:'Cormorant Garamond',serif;font-weight:700;font-size:2rem;color:var(--text);line-height:1.2;margin-bottom:0.5rem}
.wp-milestone{display:inline-block;font-size:0.62rem;padding:0.3rem 0.8rem;border-radius:6px;background:var(--blue-bg);border:1px solid var(--blue-border);color:var(--blue);margin-bottom:0.5rem}
.wp-deliv-badge{display:inline-block;font-size:0.62rem;padding:0.3rem 0.8rem;border-radius:6px;background:var(--green-bg);border:1px solid var(--green-border);color:var(--green);margin-bottom:0.5rem;margin-left:0.4rem}
.divider{height:1px;background:var(--border);margin-bottom:1.5rem}
.section-title{font-family:'Cormorant Garamond',serif;font-weight:600;font-size:1rem;color:var(--text-dim);text-transform:uppercase;letter-spacing:0.15em;margin-bottom:0.3rem}
.section-hint{font-size:0.6rem;color:var(--text-muted);margin-bottom:1.2rem;font-style:italic}
.day-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(250px,1fr));gap:0.8rem;margin-bottom:2rem}
.dc{background:var(--page);border:1px solid var(--border);border-radius:10px;padding:1.1rem 1.3rem;transition:all 0.2s;animation:cardIn 0.4s ease both;cursor:pointer}
.dc:hover{transform:translateY(-3px);box-shadow:0 6px 20px rgba(0,0,0,0.07);border-color:var(--border-light)}
.dc.holiday{cursor:default;background:var(--holiday-bg);border-color:rgba(217,119,6,0.15)}.dc.holiday:hover{transform:none;box-shadow:none}
@keyframes cardIn{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}
.dc-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:0.5rem}
.dc-day{font-family:'Cormorant Garamond',serif;font-weight:700;font-size:0.9rem;color:var(--text)}
.dc-count{font-size:0.55rem;color:var(--text-muted)}.dc-preview{font-size:0.68rem;color:var(--text-dim);line-height:1.5;margin-bottom:0.3rem}
.dc-badge{font-size:0.5rem;padding:0.12rem 0.4rem;border-radius:8px}.holiday-badge{background:var(--holiday-bg);color:var(--holiday);border:1px solid rgba(217,119,6,0.2)}
.dc-open{font-size:0.58rem;color:var(--accent);opacity:0;transition:opacity 0.2s}.dc:hover .dc-open{opacity:1}
.topic-list{display:flex;flex-direction:column;gap:0.5rem;margin-bottom:2rem}
.tl-item{background:var(--purple-bg);border:1px solid var(--purple-border);border-radius:8px;padding:0.8rem 1rem}
.tl-title{font-family:'Cormorant Garamond',serif;font-weight:700;font-size:0.9rem;color:var(--text);margin-bottom:0.2rem}
.tl-meta{display:flex;gap:0.7rem;flex-wrap:wrap;font-size:0.6rem}
.tl-inst{color:var(--purple)}.tl-time{color:var(--text-muted)}.tl-note{color:var(--text-muted);font-style:italic}
.dp-header{margin-bottom:1.8rem}.dp-meta{font-size:0.62rem;color:var(--accent);text-transform:uppercase;letter-spacing:0.15em;margin-bottom:0.4rem;font-weight:500}
.dp-title{font-family:'Cormorant Garamond',serif;font-weight:700;font-size:2.2rem;color:var(--text);line-height:1.15}
.holiday-banner{text-align:center;padding:2.5rem;background:var(--holiday-bg);border:1px solid rgba(217,119,6,0.15);border-radius:12px;font-family:'Cormorant Garamond',serif;font-size:1.2rem;color:var(--holiday);margin-bottom:2rem}
.ev-list{display:flex;flex-direction:column;gap:0.6rem;margin-bottom:2rem}
.ev-card{background:var(--page);border:1px solid var(--border);border-radius:10px;padding:1rem 1.2rem 1rem 1.5rem;display:flex;position:relative;overflow:hidden;animation:cardIn 0.35s ease both}
.ev-bar{position:absolute;left:0;top:0;width:4px;height:100%}
.ev-blue .ev-bar{background:var(--blue)}.ev-blue{background:var(--blue-bg)}
.ev-green .ev-bar{background:var(--green)}.ev-green{background:var(--green-bg)}
.ev-purple .ev-bar{background:var(--purple)}.ev-purple{background:var(--purple-bg)}
.ev-content{flex:1}
.ev-meta{display:flex;align-items:center;gap:0.5rem;flex-wrap:wrap;margin-bottom:0.35rem}
.ev-time{font-size:0.72rem;font-weight:500;color:var(--text-dim);letter-spacing:0.02em}
.ev-who{font-size:0.66rem;font-weight:500;padding:0.1rem 0.5rem;border-radius:20px;white-space:nowrap}
.ev-blue .ev-who{background:rgba(37,99,235,0.1);color:var(--blue)}
.ev-green .ev-who{background:rgba(22,163,74,0.1);color:var(--green)}
.ev-purple .ev-who{background:rgba(124,58,237,0.12);color:var(--purple)}
.ev-label{font-size:0.6rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:0.25rem}
.ev-value{font-family:'Cormorant Garamond',serif;font-weight:600;font-size:1rem;color:var(--text);line-height:1.4;white-space:pre-wrap}
.no-events{text-align:center;padding:2rem;color:var(--text-muted);font-style:italic;font-size:0.8rem}
.wp-nav{display:flex;justify-content:space-between;padding-top:1.5rem;border-top:1px solid var(--border)}
.nav-btn,.back-btn{font-family:'JetBrains Mono',monospace;font-size:0.65rem;color:var(--text-dim);background:none;border:1px solid var(--border);border-radius:6px;padding:0.4rem 1rem;cursor:pointer;transition:all 0.2s}
.nav-btn:hover,.back-btn:hover{color:var(--accent);border-color:var(--accent);background:var(--accent-glow)}
.back-btn{margin-top:1.5rem}
.footer{margin-top:3rem;padding-top:1.2rem;border-top:1px solid var(--border);text-align:center;font-size:0.58rem;color:var(--text-muted)}
@media(max-width:700px){.book{padding:2rem 1.2rem}.cover h1{font-size:1.8rem}.day-grid{grid-template-columns:1fr}}
"""

JS_TEMPLATE = """
const dayMap = __DAY_MAP__;

function go(id){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  const t=document.getElementById(id);
  if(t){requestAnimationFrame(()=>{t.classList.add('active');window.scrollTo({top:0,behavior:'smooth'})})}
}

function injectDayNav(){
  const bs='font-family:"JetBrains Mono",monospace;font-size:0.65rem;color:var(--text-dim);background:none;border:1px solid var(--border);border-radius:6px;padding:0.4rem 1rem;cursor:pointer;transition:all 0.2s';
  const hov='onmouseover="this.style.color=\\'var(--accent)\\';this.style.borderColor=\\'var(--accent)\\'" onmouseout="this.style.color=\\'var(--text-dim)\\';this.style.borderColor=\\'var(--border)\\'"';
  Object.entries(dayMap).forEach(([id,info])=>{
    let page=document.getElementById(id);
    if(!page){
      page=document.createElement('div');
      page.className='page';
      page.id=id;
      const wn=id.match(/day-(\\d+)/)[1];
      page.innerHTML='<div class="crumbs"><span class="crumb" onclick="go(\\'toc\\')">Contents</span><span class="crumb-sep">/</span><span class="crumb" onclick="go(\\''+info.week+'\\')">Week '+wn+'</span><span class="crumb-sep">/</span><span class="crumb-current">'+info.label+'</span></div><div class="dp-header"><div class="dp-meta">Week '+wn+' &middot; '+info.label+'</div><h2 class="dp-title">'+info.label+'</h2></div><div class="divider"></div><div style="text-align:center;padding:2rem;color:var(--text-muted);font-style:italic;font-size:0.8rem">'+(info.holiday||'No scheduled events')+'</div>';
      document.querySelector('.book').appendChild(page);
    }
    page.querySelectorAll('.back-btn,.day-nav').forEach(b=>b.remove());
    const nav=document.createElement('div');
    nav.className='day-nav';
    nav.style.cssText='display:flex;justify-content:space-between;align-items:center;padding-top:1.5rem;border-top:1px solid var(--border);margin-top:1rem';
    const prev=info.prev?'<button style="'+bs+'" '+hov+' onclick="go(\\''+info.prev+'\\')">\\u2190 '+dayMap[info.prev].label+'</button>':'<button style="'+bs+'" '+hov+' onclick="go(\\''+info.week+'\\')">\\u2190 Back to Week</button>';
    const mid='<button style="'+bs+'" '+hov+' onclick="go(\\''+info.week+'\\')">&#9776; Week view</button>';
    const next=info.next?'<button style="'+bs+'" '+hov+' onclick="go(\\''+info.next+'\\')">' +dayMap[info.next].label+' \\u2192</button>':'<button style="'+bs+'" '+hov+' onclick="go(\\''+info.week+'\\')">Back to Week \\u2192</button>';
    nav.innerHTML=prev+mid+next;
    page.appendChild(nav);
  });
}

document.addEventListener('DOMContentLoaded',()=>{injectDayNav();go('toc');});
"""

def gen_html(master_weeks, week_data, generated_by=None):
    today = datetime.date.today().strftime("%B %d, %Y")
    now_time = datetime.datetime.now().strftime("%I:%M %p")

    # Count stats
    total_topics = sum(
        len(w["topics"]) + sum(len(d["events"]) for d in w["days"])
        for w in week_data
    )
    total_milestones = sum(1 for m in master_weeks if m["milestone"])

    out = []
    out.append(f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{TITLE} 2026</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,400;0,600;0,700;1,400&family=JetBrains+Mono:wght@300;400;500&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>
<div class="book">
""")

    # ------------------------------------------------------------------ TOC
    out.append('<div class="page" id="toc">\n')
    out.append(f"""  <div class="cover">
    <div class="cover-orn">&loz; &loz; &loz;</div>
    <h1>{TITLE}<br>Summer 2026</h1>
    <div class="sub"><a href="https://aaec.vt.edu/academics/undergraduate/dspg.html" target="_blank" style="color:inherit;text-decoration:none">Data Science for the Public Good</a> &middot; <a href="https://www.vt.edu/" target="_blank" style="color:inherit;text-decoration:none">Virginia Tech</a></div>
    <div class="dates">{DATE_RANGE}</div>
  </div>
  <div class="toc-stats">
    <div class="ts"><div class="ts-val" style="color:var(--accent)">{len(master_weeks)}</div><div class="ts-lbl" style="background:var(--accent-glow);border:1px solid var(--accent-border);color:var(--accent);padding:0.15rem 0.6rem;border-radius:6px;letter-spacing:0.08em">&#128197; Weeks</div></div>
    <div class="ts"><div class="ts-val" style="color:var(--blue)">{total_milestones}</div><div class="ts-lbl" style="background:var(--blue-bg);border:1px solid var(--blue-border);color:var(--blue);padding:0.15rem 0.6rem;border-radius:6px;letter-spacing:0.08em">&#127937; Milestones</div></div>
    <div class="ts"><div class="ts-val" style="color:var(--green)">{sum(1 for m in master_weeks if m['deliverable'])}</div><div class="ts-lbl" style="background:var(--green-bg);border:1px solid var(--green-border);color:var(--green);padding:0.15rem 0.6rem;border-radius:6px;letter-spacing:0.08em">&#9989; Deliverables</div></div>
  </div>
  <div class="toc-heading">Table of Contents</div>
""")

    for i, mw in enumerate(master_weeks):
        wn = mw["num"]
        delay = 0.15 + i * 0.05
        wd = week_data[i] if i < len(week_data) else {}
        all_days = wd.get("days", [])
        n_days = len([d for d in all_days if not d.get("holiday")])

        # Date range from first to last day in the sheet
        date_range_str = ""
        if all_days:
            first = all_days[0]["date"]
            last  = all_days[-1]["date"]
            date_range_str = (f"{h(first)} &ndash; {h(last)} &middot; "
                              if first != last else f"{h(first)} &middot; ")

        days_str = f"{n_days} day{'s' if n_days != 1 else ''}" if n_days else ""

        out.append(f"""  <div class="chapter" onclick="go('week-{wn}')" style="animation-delay:{delay}s">
    <div class="ch-left">
      <div class="ch-num">Week {wn}</div>
      <div class="ch-title">{h(mw["topic"])}</div>
      <div class="ch-meta">{date_range_str}{days_str}</div>
    </div>
    <div class="ch-right"><div class="ch-arrow">&rarr;</div></div>
  </div>
""")

    out.append("</div>\n\n")

    # ----------------------------------------------------------- Week pages
    for i, (mw, wd) in enumerate(zip(master_weeks, week_data)):
        wn = mw["num"]
        prev_wn = master_weeks[i-1]["num"] if i > 0 else None
        next_wn = master_weeks[i+1]["num"] if i < len(master_weeks)-1 else None
        nav_prev = (f'<button class="nav-btn" onclick="go(\'week-{prev_wn}\')">&larr; Week {prev_wn}</button>'
                    if prev_wn else f'<button class="nav-btn" onclick="go(\'toc\')">&larr; Contents</button>')
        nav_next = (f'<button class="nav-btn" onclick="go(\'week-{next_wn}\')">Week {next_wn} &rarr;</button>'
                    if next_wn else '<button class="nav-btn" onclick="go(\'toc\')">Contents &#8617;</button>')

        milestone_badge = (f'<span class="wp-milestone">&#127937; {h(mw["milestone"])}</span>'
                           if mw["milestone"] else "")
        deliv_badge = (f'<span class="wp-deliv-badge">&#9989; {h(mw["deliverable"])}</span>'
                       if mw["deliverable"] else "")

        out.append(f'<div class="page" id="week-{wn}">\n')
        out.append(f'  <div class="crumbs"><span class="crumb" onclick="go(\'toc\')">Contents</span>'
                   f'<span class="crumb-sep">/</span><span class="crumb-current">Week {wn}</span></div>\n')
        out.append(f'  <div class="wp-header">\n'
                   f'    <div class="wp-label">Week {wn}</div>\n'
                   f'    <h2 class="wp-title">{h(mw["topic"])}</h2>\n'
                   f'    {milestone_badge}{deliv_badge}\n'
                   f'  </div>\n')
        out.append('  <div class="divider"></div>\n')

        days = wd.get("days", [])
        topics = wd.get("topics", [])
        week_delivs = wd.get("week_deliverables", [])

        if days:
            out.append('  <div class="section-title">Days</div>\n')
            out.append('  <div class="section-hint">Click a day to see full schedule</div>\n')
            out.append('  <div class="day-grid">\n')
            for di, day in enumerate(days):
                day_id = f"day-{wn}-{di}"
                n_ev = len(day["events"])
                if day["holiday"]:
                    out.append(f'    <div class="dc holiday">\n'
                               f'      <div class="dc-head"><div class="dc-day">{h(day["date"])}</div>'
                               f'<span class="dc-badge holiday-badge">Holiday</span></div>\n'
                               f'      <div class="dc-preview">{h(day["title"])}</div>\n'
                               f'    </div>\n')
                else:
                    out.append(f'    <div class="dc" onclick="go(\'{day_id}\')">\n'
                               f'      <div class="dc-head"><div class="dc-day">{h(day["date"])}</div>'
                               f'<div class="dc-count">{n_ev} event{"s" if n_ev != 1 else ""}</div></div>\n'
                               f'      <div class="dc-preview">{h(day["title"])}</div>\n'
                               f'      <div class="dc-open">Open &rarr;</div>\n'
                               f'    </div>\n')
            out.append('  </div>\n')

        if topics:
            out.append('  <div class="section-title">Topics</div>\n')
            out.append('  <div class="topic-list" style="margin-top:0.8rem">\n')
            for t in topics:
                inst_html = f'<span class="tl-inst">{h(t["instructor"])}</span>' if t["instructor"] else ""
                time_html = f'<span class="tl-time">{h(t["time"])}</span>' if t["time"] else ""
                note_html = f'<span class="tl-note">{h(t["note"])}</span>' if t["note"] else ""
                out.append(f'    <div class="tl-item">\n'
                           f'      <div class="tl-title">{h(t["title"])}</div>\n'
                           f'      <div class="tl-meta">{inst_html}{time_html}{note_html}</div>\n'
                           f'    </div>\n')
            out.append('  </div>\n')

        if week_delivs:
            out.append('  <div class="divider"></div>\n')
            out.append('  <div class="section-title">Week Deliverables</div>\n')
            out.append('  <div class="ev-list" style="margin-top:0.8rem">\n')
            for dv in week_delivs:
                out.append(ev_card_html(dv, None, None, color="ev-green"))
            out.append('  </div>\n')

        out.append(f'  <div class="wp-nav">{nav_prev}{nav_next}</div>\n')
        out.append('</div>\n\n')

        # --------------------------------------------------- Day detail pages
        for di, day in enumerate(days):
            if day["holiday"]:
                continue
            day_id = f"day-{wn}-{di}"
            out.append(f'<div class="page" id="{day_id}">\n')
            out.append(f'  <div class="crumbs">'
                       f'<span class="crumb" onclick="go(\'toc\')">Contents</span>'
                       f'<span class="crumb-sep">/</span>'
                       f'<span class="crumb" onclick="go(\'week-{wn}\')">Week {wn}</span>'
                       f'<span class="crumb-sep">/</span>'
                       f'<span class="crumb-current">{h(day["date"])}</span></div>\n')
            out.append(f'  <div class="dp-header">'
                       f'<div class="dp-meta">Week {wn} &middot; {h(day["date"])}</div>'
                       f'<h2 class="dp-title">{h(day["title"])}</h2></div>\n')
            out.append('  <div class="divider"></div>\n')

            if day["events"]:
                out.append('  <div class="ev-list">\n')
                for ei, ev in enumerate(day["events"]):
                    out.append("  " + ev_card_html(ev["activity"], ev["time"], ev["who"],
                                                   color=ev["color"], delay=(ei+1)*0.05))
                out.append('  </div>\n')
            else:
                out.append('  <div class="ev-list"><div class="no-events">No events recorded for this day.</div></div>\n')

            out.append('</div>\n\n')

    # Build dayMap for navigation
    import json
    day_entries = {}
    for i, (mw, wd) in enumerate(zip(master_weeks, week_data)):
        wn = mw["num"]
        days = wd.get("days", [])
        day_ids = [f"day-{wn}-{di}" for di in range(len(days))]
        for di, day in enumerate(days):
            day_id = f"day-{wn}-{di}"
            entry = {
                "label": day["date"],
                "week": f"week-{wn}",
                "prev": day_ids[di-1] if di > 0 else None,
                "next": day_ids[di+1] if di < len(days)-1 else None,
            }
            if day["holiday"]:
                entry["holiday"] = h(day["title"])
            day_entries[day_id] = entry

    js = JS_TEMPLATE.replace("__DAY_MAP__", json.dumps(day_entries))

    by_str = f" &middot; {h(generated_by)}" if generated_by else ""
    out.append(f'<div class="footer">Auto-generated from curriculum xlsx &middot; DSPG 2026 &middot; Virginia Tech &middot; Generated {today}, {now_time}{by_str}</div>\n')
    out.append(f'<script>{js}</script>\n')
    out.append('</div>\n</body>\n</html>\n')

    return "".join(out)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    generated_by = None
    if "--by" in args:
        idx = args.index("--by")
        if idx + 1 < len(args):
            generated_by = args[idx + 1]
        args = [a for i,a in enumerate(args) if a != "--by" and (i == 0 or args[i-1] != "--by")]
    xlsx_path = Path(args[0]) if len(args) >= 1 else Path(DEFAULT_XLSX)
    html_path = Path(args[1]) if len(args) >= 2 else Path(DEFAULT_HTML)
    # Optional: --by "Name"
    generated_by = None
    if "--by" in args:
        idx = args.index("--by")
        if idx + 1 < len(args):
            generated_by = args[idx + 1]

    if not xlsx_path.exists():
        print(f"ERROR: Cannot find '{xlsx_path}'")
        print(f"Usage: python generate_curriculum.py [input.xlsx] [output.html]")
        sys.exit(1)

    print(f"Reading  : {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Master sheet
    if "Master" not in wb.sheetnames:
        print("ERROR: No 'Master' sheet found in workbook.")
        sys.exit(1)
    master_weeks = parse_master(wb["Master"])
    print(f"  Master : {len(master_weeks)} weeks found")

    # Week sheets
    week_data = []
    for mw in master_weeks:
        sheet_name = f"Week {mw['num']}"
        # Try with trailing space too (Excel quirk)
        candidates = [sheet_name, sheet_name + " "]
        ws = None
        for c in candidates:
            if c in wb.sheetnames:
                ws = wb[c]
                break
        if ws is None:
            print(f"  Week {mw['num']}: sheet not found — skipping")
            week_data.append({"title": "", "has_days": False, "days": [], "topics": [], "week_deliverables": []})
            continue
        parsed = parse_week_sheet(ws)
        n_days   = len(parsed["days"])
        n_topics = len(parsed["topics"])
        print(f"  Week {mw['num']}: {n_days} days, {n_topics} topics, {len(parsed['week_deliverables'])} deliverables")
        week_data.append(parsed)

    html = gen_html(master_weeks, week_data, generated_by=generated_by)

    html_path.write_text(html, encoding="utf-8")
    print(f"\nGenerated: {html_path}  ({len(html):,} bytes)")
    print("Done ✓")

if __name__ == "__main__":
    main()
